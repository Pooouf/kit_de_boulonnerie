# %%
import pandas as pd
import tkinter as tk
from tkinter.messagebox import showinfo, showerror
from tkinter.filedialog import askopenfilename
from tkinter import ttk
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl import load_workbook, workbook, worksheet
from copy import copy
import pathlib
import os

os.path.realpath(__file__)

# %%
type_gtr = 'CFM56-5B'
module = '03X'
cas_de_demontage = '000'
DATA_FILE = 'Préparation Kits Boulonnerie U3_ (Récupéré).xlsm'

data_filepath = pathlib.Path(__file__).parent / DATA_FILE
data_filename = askopenfilename(
    title=f'Ouvrir le fichier modèle d\'impression "{DATA_FILE}"',
#    message="Note : Le fichier doit être un tableur Excel, contenir un onglet CFM56-5B et un onglet Impression",
    initialfile=data_filepath,
)

data = pd.read_excel(data_filename, sheet_name='CFM56-5B', header=1)

# %%
module_names = sorted(data['Module'].drop_duplicates())

# %%
column_names = pd.read_excel(DATA_FILE, sheet_name='CFM56-5B', header=0, nrows=1)
premier_cas_de_demontage = list(column_names.columns).index('Cas de démontage')
columns = list(data.columns)
cas_de_demontages = sorted(columns[premier_cas_de_demontage:])

# %%
root = tk.Tk()
screen_width = root.winfo_screenwidth()
screen_height = root.winfo_screenheight()
window_width = 300
window_height = 520
x = int(screen_width/2 - window_width/2)
y = int(screen_height/2 - window_height/2)
root.geometry(f'{window_width}x{window_height}+{x}+{y}')
root.title("Préparation des Kits Boulonnerie U3")

text = ttk.Label(root, text="Choisissez un module :")
text.pack()

var_module=tk.IntVar()
for index, module_name in enumerate(module_names):
    button = ttk.Radiobutton(root, text=module_name, value=index, variable=var_module)
    button.pack()


text = ttk.Label(root, text="Choisissez un cas de démontage :")
text.pack()

var_demontage=tk.IntVar()
for index, name in enumerate(cas_de_demontages):
    button = ttk.Radiobutton(root, text=name, value=index, variable=var_demontage)
    button.pack()


button = ttk.Button(text="Générer les kits", command=lambda: root.destroy())
button.pack()

# Start the event loop
root.mainloop()

module = module_names[var_module.get()]
cas_de_demontage = cas_de_demontages[var_demontage.get()]

print('type_gtr :', type_gtr)
print('module :', module)
print('cas_de_demontage :', cas_de_demontage)


# %%
# debug lines
# colmns_to_remove = [cas for cas in cas_de_demontages if cas != cas_de_demontage]
# columns_of_interest = [column for column in data.columns if column not in colmns_to_remove]
# kit = data[columns_of_interest]
# kit = kit[(kit['Module'] == module) & (kit[cas_de_demontage] == 'x')]

# non-debug lines
columns_of_interest = [column for column in data.columns if column not in cas_de_demontages]
kits = data[(data['Module'] == module) & (data[cas_de_demontage] == 'x')]
kits = kits[columns_of_interest]


# %%
TEMPLATE_NAME = 'Impression'
OUTPUT_FILENAME = 'impression.xlsx'

wb = load_workbook(filename=DATA_FILE)
for sheetname in wb.sheetnames:
    if sheetname == TEMPLATE_NAME:
        continue
    del wb[sheetname]
template = wb[TEMPLATE_NAME]
template['B4'].value = type_gtr
template['B6'].value = module
template['B8'].value = cas_de_demontage

# les lignes 15 et 16 (en notation 1-based) contiennent des valeurs. Supprimons-les avant d'ajouter les bonnes.
# template['']
INSERTION_IDX = 15
emphasize_font = copy(template[f'A{INSERTION_IDX}'].font)
regular_font = copy(template[f'B{INSERTION_IDX}'].font)
template.delete_rows(INSERTION_IDX, 2)

for kit_name in kits['Libellé Kit'].drop_duplicates():
    sheet = wb.copy_worksheet(template)
    sheet.title = kit_name.replace('/', '-')[:31]
    sheet['B11'].value = kit_name

    kit = kits[kits['Libellé Kit'] == kit_name].filter(['PN', 'Libellé', 'PN Alt.', 'Qté', 'Neuf', 'Conditions'])
    row_count = kit.shape[0]
    sheet.insert_rows(INSERTION_IDX, row_count)
    for row in dataframe_to_rows(kit, index=False, header=False):
        sheet.append(row)
    top_idx = sheet.max_row - row_count +1
    sheet.move_range(f"A{top_idx}:F{top_idx+row_count}", rows=INSERTION_IDX-top_idx, cols=0)

    for i in range(INSERTION_IDX, INSERTION_IDX+row_count):
        sheet[f'A{i}'].font = emphasize_font
        sheet[f'B{i}'].font = regular_font
        sheet[f'C{i}'].font = regular_font
        sheet[f'D{i}'].font = emphasize_font
        sheet[f'E{i}'].font = regular_font
        sheet[f'F{i}'].font = regular_font
        sheet[f'G{i}'].font = regular_font
        sheet[f'H{i}'].font = regular_font

del wb[TEMPLATE_NAME]

try:
    wb.save(OUTPUT_FILENAME)
    showinfo(
        title="Préparation des Kits Boulonnerie U3",
        message=f'Les kits ont été enregistrés dans le fichier {OUTPUT_FILENAME}'
    )
except Exception as exc:
    showerror(
        title="Préparation des Kits Boulonnerie U3",
        message=f'''
            Type : {type_gtr}
            Module : {module}
            Cas de demontage : {cas_de_demontage}
            Modèle d'impression : {data_filename}
            Erreur : {str(exc)}
        '''
    )
